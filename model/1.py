from openpyxl import load_workbook
import random
from numpy import *
from PIL import Image
import os
workbook1 = load_workbook(r'D:\下载\2021_MCM_Problem_C_Data\2021_MCM_Problem_C_Data\2021MCMProblemC_DataSet.xlsx')    #找到需要xlsx文件的位置
MCMProblemC_DataSet = workbook1.active                 #获取当前活跃的sheet,默认是第一个sheet


#如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
# sheets = workbook1.get_sheet_names()  # 从名称获取sheet
# MCMProblemC_DataSet = workbook1.get_sheet_by_name(sheets[0])

#获取sheet页的行数据
rows = MCMProblemC_DataSet.rows
#获取sheet页的列数据
columns = MCMProblemC_DataSet.columns

# feature = open('feature.txt', mode='w')#只读。
# a = str(MCMProblemC_DataSet.cell(row=1, column=2).value) # 将excel中 1行2列 对应的数据传给a
sizerow=MCMProblemC_DataSet.max_row         #读取excel行数
sizecol=MCMProblemC_DataSet.max_column     #读取excel列数
print(sizerow,sizecol)

positive=[]
negative=[]
unverified=[]
unprocessed=[]

i = 0
a=0
b=0
c=0
d = 0

head = MCMProblemC_DataSet.cell(row=1, column=1).value
for i in range(2, sizerow + 1):
    # print(head)
    # if head == 'GlobalID':
    #     print(j,"GlobalID")
    # if head == 'Lab Status':
    #     print(j,"Lab Status")
    GlobalID=MCMProblemC_DataSet.cell(row=i, column=1).value
    Status=MCMProblemC_DataSet.cell(row=i, column=4).value
    Latitude = MCMProblemC_DataSet.cell(row=i, column=7).value
    Longitude = MCMProblemC_DataSet.cell(row=i, column=8).value

    if Status=='Positive ID':
        positive.append([Latitude,Longitude])
        a=a+1
        continue
    if Status == 'Negative ID':
        negative.append([Latitude, Longitude])
        b=b+1
        continue
    if Status == 'Unverified':
        unverified.append([Latitude, Longitude])
        c=c+1
        continue
    if Status == 'Unprocessed':
        unprocessed.append([Latitude, Longitude])
        d=d+1
meanpositive =  mean(positive, axis = 0)
meannegative =  mean(negative, axis = 0)

print(a,b,c,d)
print(meanpositive,meannegative)
name={}

def euclDistance(vector1, vector2):
    return sqrt(sum(power(vector2 - vector1, 2)))  #求这两个矩阵的距离，vector1、2均为矩阵

c=0
d=0
for i in range(2, sizerow + 1):
    # print(head)
    # if head == 'GlobalID':
    #     print(j,"GlobalID")
    # if head == 'Lab Status':
    #     print(j,"Lab Status")
    GlobalID=MCMProblemC_DataSet.cell(row=i, column=1).value
    Status=MCMProblemC_DataSet.cell(row=i, column=4).value
    Latitude = MCMProblemC_DataSet.cell(row=i, column=7).value
    Longitude = MCMProblemC_DataSet.cell(row=i, column=8).value
    ll=[Latitude,Latitude]
    if Status == 'Unverified':
        print(euclDistance(ll,meanpositive),euclDistance(ll,meannegative))
        if  euclDistance(ll,meanpositive) <= euclDistance(ll,meannegative):
            name[GlobalID]=1
            c=c+1
        else:
            name[GlobalID] = 0
            d=d+1

    if Status == 'Unprocessed':
        if euclDistance(ll,meanpositive)<=euclDistance(ll,meannegative):
                name[GlobalID]=1
                c = c + 1
        else:
            name[GlobalID] = 0
            d = d + 1
print(name)
print(c,d)