from openpyxl import load_workbook
import random

from PIL import Image
import os
workbook1 = load_workbook(r'D:\下载\2021_MCM_Problem_C_Data\2021_MCM_Problem_C_Data\2021MCMProblemC_DataSet.xlsx')    #找到需要xlsx文件的位置
MCMProblemC_DataSet = workbook1.active                 #获取当前活跃的sheet,默认是第一个sheet


#如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
# sheets = workbook1.get_sheet_names()  # 从名称获取sheet
# MCMProblemC_DataSet = workbook1.get_sheet_by_name(sheets[0])

#获取sheet页的行数据
rows = MCMProblemC_DataSet.rows
#获取sheet页的列数据
columns = MCMProblemC_DataSet.columns

# feature = open('feature.txt', mode='w')#只读。
# a = str(MCMProblemC_DataSet.cell(row=1, column=2).value) # 将excel中 1行2列 对应的数据传给a
sizerow=MCMProblemC_DataSet.max_row         #读取excel行数
sizecol=MCMProblemC_DataSet.max_column     #读取excel列数
print(sizerow,sizecol)
name = {}
i = 0
a=0
b=0
c=0
d = 0

head = MCMProblemC_DataSet.cell(row=1, column=1).value
for i in range(2, sizerow + 1):
    # print(head)
    # if head == 'GlobalID':
    #     print(j,"GlobalID")
    # if head == 'Lab Status':
    #     print(j,"Lab Status")
    GlobalID=MCMProblemC_DataSet.cell(row=i, column=1).value
    Status=MCMProblemC_DataSet.cell(row=i, column=4).value
    if Status=='Positive ID':
        name[GlobalID]=1
        a=a+1
        continue
    if Status == 'Negative ID':
        name[GlobalID] = 0
        b=b+1
        continue
    if Status == 'Unverified':
        name[GlobalID] = -1
        c=c+1
        continue
    else :
        name[GlobalID] = -2
        # print(Status)
        d=d+1




        # for i in range(1,sizerow+1):
print(a,b,c,d)

workbook2 = load_workbook(r'D:\下载\2021_MCM_Problem_C_Data\2021_MCM_Problem_C_Data\2021MCM_ProblemC_ Images_by_GlobalID.xlsx')    #找到需要xlsx文件的位置
Images_by_GlobalID = workbook2.active                 #获取当前活跃的sheet,默认是第一个sheet

sizerowImages_by_GlobalID=Images_by_GlobalID.max_row         #读取excel行数
sizecolImages_by_GlobalID=Images_by_GlobalID.max_column     #读取excel列数
print(sizerowImages_by_GlobalID,sizecolImages_by_GlobalID)

lable = open('./dataset/lable.txt', mode='w')
trainlable = open('./dataset/trainlable.txt', mode='w')
vallable = open('./dataset/vallable.txt', mode='w')
testlable = open('./dataset/testlable.txt', mode='w')
unverifiedlable = open('./dataset/unverifiedlable.txt', mode='w')
line=''
trainline=''
valline=''
testline=''
n=0
k=0

m=0
l=0
w=0
e=0
r=0
valone=0
testone=0

def PNG_JPG(PngPath):
    img = cv.imread(PngPath, 0)
    w, h = img.shape[::-1]
    infile = PngPath
    outfile = os.path.splitext(infile)[0] + ".jpg"
    img = Image.open(infile)
    img = img.resize((int(w / 2), int(h / 2)), Image.ANTIALIAS)
    try:
        if len(img.split()) == 4:
            # prevent IOError: cannot write mode RGBA as BMP
            r, g, b, a = img.split()
            img = Image.merge("RGB", (r, g, b))
            img.convert('RGB').save(outfile, quality=70)
            os.remove(PngPath)
        else:
            img.convert('RGB').save(outfile, quality=70)
            os.remove(PngPath)
        return outfile
    except Exception as e:
        print("PNG转换JPG  错误", e)


for j in range(2,sizerowImages_by_GlobalID+1):
    GlobalID = Images_by_GlobalID.cell(row=j, column=2).value
    FileName = Images_by_GlobalID.cell(row=j, column=1).value
    FileType = Images_by_GlobalID.cell(row=j, column=3).value
    # print(FileType)
    if GlobalID in name.keys():#如果在
        if FileType == 'image/jpg' or FileType == 'image/png' :
            if FileType == 'image/png' :
                FileName = os.path.splitext(FileName)[0] + ".jpg"

            if name[GlobalID]==-1 or name[GlobalID] ==-2:
                unverifiedlable.write(FileName+','+str(0)+ '\n')

            if name[GlobalID]==0 or name[GlobalID] ==1:
                line = FileName + ','
                line = line+ str(name[GlobalID])+ '\n'
                lable.write(line)
                e=e+1

                # print(line)
                k=random.randint(0,100)
                n = random.randint(0,100)
                aa= random.randint(0,100)

                if name[GlobalID]==1:
                    trainlable.write(line)
                    vallable.write(line)
                    testlable.write(line)
                    r=r+1

                if name[GlobalID]==0:
                    if k<= 30 and l<30:
                        vallable.write(line)
                        l=l+1
                        continue
                    if 30< n<=60 and m<30:
                        testlable.write(line)
                        m=m+1
                        continue
                    if 60< aa<=90 and w<=400:
                        trainlable.write(line)
                        w=w+1
                        continue



print(r,l,m,w,e)

trainlable.close()
testlable.close()
vallable.close()
lable.close()

lable = open('./dataset/lable.txt', mode='a')
trainlable = open('./dataset/trainlable.txt', mode='a')
vallable = open('./dataset/vallable.txt', mode='a')
testlable = open('./dataset/testlable.txt', mode='a')
i = 0
line=''
for i in range(133):
    line = str(i) + '.jpg'+',' + '1' + '\n'
    cc = random.randint(0, 100)
    if 0<cc<=20 and valone<16:
        vallable.write(line)
        valone=valone+1
    if 20<cc<=40 and testone<16:
        testlable.write(line)
        testone=testone+1

    trainlable.write(line)
    lable.write(line)


trainlable.close()
testlable.close()
vallable.close()
lable.close()
unverifiedlable.close()



# for column in columns:# 迭代所有的列
#     for row in rows:# 迭代所有的行
#             if column.value=='GlobalID':
#             i = i + 1
#             # line = [col.value for col in row]
#             # print(line)
#             cell_data_1 = MCMProblemC_DataSet.cell(row=i, column=3).value               #获取第i行1 列的数据

